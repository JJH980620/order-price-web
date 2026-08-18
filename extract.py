# -*- coding: utf-8 -*-
"""从 订单价格申请表6.4.xlsx 提取全部数据源，生成 data.js"""
import openpyxl, json, os

SRC = r"C:\Users\admin\Desktop\订单价格申请表6.4.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.js")

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------- 1月政策一览表：车系 -> 车型(指导价/成本价) ----------
pol = wb["1月政策一览表"]
vehicles = {}
order = []
for r in range(3, pol.max_row + 1):
    car = pol.cell(row=r, column=1).value
    model = pol.cell(row=r, column=2).value
    guide = pol.cell(row=r, column=3).value
    cost = pol.cell(row=r, column=7).value
    if car and model and guide is not None:
        if car not in vehicles:
            vehicles[car] = []
            order.append(car)
        vehicles[car].append({"model": model, "guide": guide, "cost": cost})

# ---------- 装潢保养价格表 ----------
fit = wb["装潢保养价格表"]
# 各车系 -> 精品(名称, 采购价)
def gather(range_rows):
    return {car: [] for car in order}

jingpin = {car: [] for car in order}
for r in range(2, fit.max_row + 1):
    car = fit.cell(row=r, column=1).value
    name = fit.cell(row=r, column=3).value
    buy = fit.cell(row=r, column=4).value
    if car in jingpin and name and buy is not None:
        if not any(x["name"] == name for x in jingpin[car]):
            jingpin[car].append({"name": name, "buy": round(buy, 2)})

# 颜色加价（主表 B5 硬编码逻辑，作为初始迁移基准）
colorPremium = {"霜雾灰": 3000, "青峦翠": 3000, "云夜紫": 5000, "赤沙红": 5000,
                "星月银/悬浮顶": 10000, "破晓金/悬浮顶": 10000, "湖光绿/悬浮顶": 10000}

# 车身颜色列映射（车系 -> Excel 列号），每种颜色带该车系的加价
color_col = {"钛3": 19, "钛3闪充版": 20, "钛7": 21, "豹5": 22, "豹5智驾版": 23,
             "豹5长续航版": 24, "豹8": 25, "钛7闪充版": 26, "豹5闪充版": 29, "豹8闪充版": 31}
colors = {car: [] for car in order}
for car, col in color_col.items():
    if car not in colors:
        continue
    seen = set()
    for r in range(2, fit.max_row + 1):
        v = fit.cell(row=r, column=col).value
        if v and v not in seen:
            seen.add(v)
            colors[car].append({"name": v, "premium": colorPremium.get(v, 0)})

# 内饰列映射
interior_col = {"钛3": 33, "钛3闪充版": 34, "钛7": 35, "豹5": 36, "豹5智驾版": 37,
                "豹5长续航版": 38, "豹8": 39, "钛7闪充版": 40, "豹5闪充版": 42, "豹8闪充版": 43}
interiors = {car: [] for car in order}
for car, col in interior_col.items():
    if car not in interiors:
        continue
    seen = set()
    for r in range(2, fit.max_row + 1):
        v = fit.cell(row=r, column=col).value
        if v and v not in seen:
            seen.add(v)
            interiors[car].append(v)

# 保养价格 G13:I23
maintain = {}
for r in range(14, 24):
    car = fit.cell(row=r, column=7).value
    sell = fit.cell(row=r, column=8).value
    cost = fit.cell(row=r, column=9).value
    if car and sell is not None and cost is not None:
        maintain[car] = {"sell": sell, "cost": cost}

# ---------- 贷款-保险 ----------
ln = wb["贷款-保险"]
banks = []
for r in range(2, 15):
    b = ln.cell(row=r, column=1).value
    rate = ln.cell(row=r, column=4).value
    if b and rate is not None:
        banks.append({"bank": b, "rate": rate})
loan = {}
for r in range(2, 12):
    car = ln.cell(row=r, column=7).value
    ins = ln.cell(row=r, column=8).value
    limit = ln.cell(row=r, column=9).value
    if car:
        loan[car] = {"insBonus": ins, "limit": limit}

data = {
    "vehicles": vehicles,
    "carOrder": order,
    "jingpin": jingpin,
    "colors": colors,
    "interiors": interiors,
    "maintain": maintain,
    "banks": banks,
    "loan": loan,
}

with open(OUT, "w", encoding="utf-8") as f:
    f.write("// 自动生成：订单价格申请表 数据源\n")
    f.write("window.ORDER_DATA = ")
    f.write(json.dumps(data, ensure_ascii=False, indent=1))
    f.write(";\n")

print("已生成:", OUT)
print("车系列表:", order)
print("银行:", [b["bank"] for b in banks])
print("车系限价:", loan)
